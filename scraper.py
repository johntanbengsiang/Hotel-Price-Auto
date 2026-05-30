#!/usr/bin/env python3
"""
Hotel Price Scraper
-------------------
Reads configuration from the CONFIG_JSON environment variable (set via GitHub Secret).
Scrapes Booking.com for the cheapest available room per hotel/date combination.
Appends results to hotel_prices.xlsx and uploads it to OneDrive.

Environment variables required (all set as GitHub Secrets):
  CONFIG_JSON          - JSON string with hotels, dates, adults, rooms, currency
  ONEDRIVE_TOKEN       - Microsoft Graph API access token (refreshed each run)
  ONEDRIVE_CLIENT_ID   - Azure App client ID
  ONEDRIVE_CLIENT_SECRET - Azure App client secret
  ONEDRIVE_TENANT_ID   - Azure tenant ID (use "consumers" for personal OneDrive)
  ONEDRIVE_FOLDER      - Target folder path in OneDrive e.g. "Hotel Tracker"
"""

import asyncio
import io
import json
import os
import re
import sys
from dataclasses import dataclass, fields
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import requests


# ── Load config from environment ─────────────────────────

def load_config() -> dict:
    raw = os.environ.get("CONFIG_JSON", "").strip()
    if not raw:
        print("❌  CONFIG_JSON environment variable is not set.")
        print("    Set it as a GitHub Secret containing your config.json contents.")
        sys.exit(1)
    try:
        return json.loads(raw)
    except json.JSONDecodeError as e:
        print(f"❌  CONFIG_JSON is not valid JSON: {e}")
        sys.exit(1)


CONFIG       = load_config()
HOTELS       = CONFIG["hotels"]
DATE_RANGES  = [tuple(d) for d in CONFIG["date_ranges"]]
ADULTS       = CONFIG.get("adults", 2)
ROOMS        = CONFIG.get("rooms", 1)
CURRENCY     = CONFIG.get("currency", "SGD")
OUTPUT_FILE  = Path("hotel_prices.xlsx")


# ── Data model ────────────────────────────────────────────

@dataclass
class RoomResult:
    scraped_date:       str
    hotel:              str
    check_in:           str
    check_out:          str
    nights:             int
    room_type:          str
    price_per_night:    str
    total_price:        str
    currency:           str
    free_cancellation:  str
    breakfast_included: str
    url:                str


# ── Playwright helpers ────────────────────────────────────

async def dismiss_overlays(page) -> None:
    for selector in [
        'button[id*="onetrust-accept"]',
        'button[aria-label*="Dismiss"]',
        'button[aria-label*="Close"]',
        '[data-testid="user-account-ui-cta-close"]',
        'button.modal-mask-closeBtn',
    ]:
        try:
            btn = page.locator(selector).first
            if await btn.is_visible(timeout=1500):
                await btn.click()
                await page.wait_for_timeout(600)
        except Exception:
            pass


async def get_hotel_url(page, hotel: str, check_in: str, check_out: str) -> Optional[str]:
    q = hotel.replace(" ", "+")
    search_url = (
        f"https://www.booking.com/search.html?ss={q}"
        f"&checkin={check_in}&checkout={check_out}"
        f"&group_adults={ADULTS}&no_rooms={ROOMS}"
        f"&selected_currency={CURRENCY}&lang=en-gb&order=popularity"
    )
    await page.goto(search_url, wait_until="domcontentloaded", timeout=35000)
    await page.wait_for_timeout(3500)
    await dismiss_overlays(page)

    card = page.locator('[data-testid="property-card"]').first
    try:
        await card.wait_for(timeout=8000)
    except Exception:
        return None

    href = await card.locator('[data-testid="title-link"]').first.get_attribute("href")
    if not href:
        return None
    if not href.startswith("http"):
        href = "https://www.booking.com" + href
    return href.split("?")[0]


async def scrape_room_table(page) -> list[dict]:
    rooms = []
    try:
        await page.wait_for_selector("#hprt-table, .hprt-table", timeout=12000)
    except Exception:
        return rooms

    rows = await page.query_selector_all(".hprt-table tr, #hprt-table tr")
    current_room_name = "N/A"

    for row in rows:
        name_el = await row.query_selector(".hprt-roomtype-icon-link, .room-name")
        if name_el:
            t = (await name_el.inner_text()).strip()
            if t:
                current_room_name = t

        price_el = await row.query_selector(
            ".bui-price-display__value, "
            ".prco-valign-middle-helper, "
            "[data-testid='price-and-discounted-price']"
        )
        if not price_el:
            continue

        price_raw = (await price_el.inner_text()).strip()
        nums = re.findall(r"[\d,]+(?:\.\d+)?", price_raw.replace(",", ""))
        if not nums:
            continue
        try:
            total_val = float(nums[-1])
        except ValueError:
            continue

        row_text = (await row.inner_text()).lower()
        cond_el = await row.query_selector(".hprt-conditions, [data-testid='cancellation-and-charges']")
        check_cond = (await cond_el.inner_text()).lower() if cond_el else row_text

        free_cancel = "Unknown"
        if "free cancellation" in check_cond:
            free_cancel = "Yes"
        elif "non-refundable" in check_cond or "no refund" in check_cond:
            free_cancel = "No"

        breakfast = "Unknown"
        if "breakfast included" in row_text or "includes breakfast" in row_text:
            breakfast = "Yes"
        elif "room only" in row_text or "no breakfast" in row_text or "without breakfast" in row_text:
            breakfast = "No"
        elif "breakfast" in row_text:
            breakfast = "Yes"

        rooms.append({
            "room_type":         current_room_name,
            "total_val":         total_val,
            "free_cancellation": free_cancel,
            "breakfast":         breakfast,
        })

    rooms.sort(key=lambda r: r["total_val"])
    return rooms


async def fetch_hotel(page, hotel: str, check_in: str, check_out: str) -> Optional[RoomResult]:
    ci = datetime.strptime(check_in, "%Y-%m-%d")
    co = datetime.strptime(check_out, "%Y-%m-%d")
    nights = (co - ci).days

    print(f"  🔍  {hotel}  |  {check_in} → {check_out}")
    try:
        hotel_url = await get_hotel_url(page, hotel, check_in, check_out)
        if not hotel_url:
            print(f"    ⚠️  No search result found")
            return None

        direct_url = (
            f"{hotel_url}?checkin={check_in}&checkout={check_out}"
            f"&group_adults={ADULTS}&no_rooms={ROOMS}"
            f"&selected_currency={CURRENCY}&lang=en-gb"
        )
        await page.goto(direct_url, wait_until="domcontentloaded", timeout=35000)
        await page.wait_for_timeout(3500)
        await dismiss_overlays(page)

        rooms = await scrape_room_table(page)
        if not rooms:
            print(f"    ⚠️  Room table empty")
            return None

        cheapest  = rooms[0]
        total     = cheapest["total_val"]
        per_night = round(total / nights, 0)

        print(f"    ✅  {cheapest['room_type'][:40]}  |  {CURRENCY} {per_night:.0f}/night")
        return RoomResult(
            scraped_date       = date.today().isoformat(),
            hotel              = hotel,
            check_in           = check_in,
            check_out          = check_out,
            nights             = nights,
            room_type          = cheapest["room_type"],
            price_per_night    = f"{per_night:.0f}",
            total_price        = f"{total:.0f}",
            currency           = CURRENCY,
            free_cancellation  = cheapest["free_cancellation"],
            breakfast_included = cheapest["breakfast"],
            url                = direct_url,
        )
    except Exception as e:
        print(f"    ❌  Error: {e}")
        return None


# ── Excel helpers ─────────────────────────────────────────

def build_workbook(results: list[RoomResult]) -> bytes:
    """
    Download existing workbook from OneDrive (if it exists), append new rows,
    and return the updated workbook as bytes.
    """
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    col_names = [f.name for f in fields(RoomResult)]

    # Try to fetch existing file from OneDrive
    existing_bytes = download_from_onedrive()
    if existing_bytes:
        wb = load_workbook(io.BytesIO(existing_bytes))
        ws = wb.active
    else:
        # First run — create fresh workbook with styled header
        wb = Workbook()
        ws = wb.active
        ws.title = "Hotel Prices"

        header_fill = PatternFill("solid", fgColor="1F3864")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        widths = {
            "scraped_date": 14, "hotel": 30, "check_in": 12, "check_out": 12,
            "nights": 8, "room_type": 32, "price_per_night": 14,
            "total_price": 12, "currency": 10, "free_cancellation": 18,
            "breakfast_included": 18, "url": 55,
        }
        for col, name in enumerate(col_names, 1):
            cell = ws.cell(row=1, column=col,
                           value=name.replace("_", " ").title())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = widths.get(name, 15)

    # Append new rows
    next_row = ws.max_row + 1
    stripe_a = PatternFill("solid", fgColor="EBF0FA")
    stripe_b = PatternFill("solid", fgColor="FFFFFF")

    for i, r in enumerate(results):
        fill = stripe_a if (next_row + i) % 2 == 0 else stripe_b
        for col, name in enumerate(col_names, 1):
            val = getattr(r, name)
            if name in ("price_per_night", "total_price", "nights"):
                try:
                    val = int(val) if "." not in str(val) else float(val)
                except (ValueError, TypeError):
                    pass
            cell = ws.cell(row=next_row + i, column=col, value=val)
            cell.fill = fill
            cell.alignment = Alignment(
                horizontal="left" if col == 2 else "center"
            )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── OneDrive (Microsoft Graph API) ───────────────────────

def get_onedrive_token() -> str:
    """Fetch a fresh OAuth2 token using client credentials."""
    tenant   = os.environ["ONEDRIVE_TENANT_ID"]      # "consumers" for personal
    client_id     = os.environ["ONEDRIVE_CLIENT_ID"]
    client_secret = os.environ["ONEDRIVE_CLIENT_SECRET"]

    # Personal OneDrive uses the "consumers" tenant endpoint
    if tenant == "consumers":
        url = "https://login.microsoftonline.com/consumers/oauth2/v2.0/token"
    else:
        url = f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token"

    resp = requests.post(url, data={
        "grant_type":    "client_credentials",
        "client_id":     client_id,
        "client_secret": client_secret,
        "scope":         "https://graph.microsoft.com/.default",
    }, timeout=30)
    resp.raise_for_status()
    return resp.json()["access_token"]


def _graph_headers(token: str) -> dict:
    return {"Authorization": f"Bearer {token}",
            "Content-Type": "application/octet-stream"}


def download_from_onedrive() -> Optional[bytes]:
    """Return existing hotel_prices.xlsx bytes from OneDrive, or None if not found."""
    try:
        token  = get_onedrive_token()
        folder = os.environ.get("ONEDRIVE_FOLDER", "Hotel Tracker")
        path   = f"{folder}/hotel_prices.xlsx"
        url    = f"https://graph.microsoft.com/v1.0/me/drive/root:/{path}:/content"
        resp   = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=30)
        if resp.status_code == 200:
            print("  📥  Downloaded existing workbook from OneDrive")
            return resp.content
        elif resp.status_code == 404:
            print("  📄  No existing workbook found — creating fresh one")
            return None
        else:
            print(f"  ⚠️  OneDrive download returned {resp.status_code}, starting fresh")
            return None
    except Exception as e:
        print(f"  ⚠️  Could not download from OneDrive: {e}")
        return None


def upload_to_onedrive(file_bytes: bytes) -> None:
    """Upload hotel_prices.xlsx bytes to OneDrive, overwriting any existing file."""
    token  = get_onedrive_token()
    folder = os.environ.get("ONEDRIVE_FOLDER", "Hotel Tracker")
    path   = f"{folder}/hotel_prices.xlsx"
    url    = f"https://graph.microsoft.com/v1.0/me/drive/root:/{path}:/content"

    resp = requests.put(url, headers=_graph_headers(token),
                        data=file_bytes, timeout=60)
    if resp.status_code in (200, 201):
        print(f"  ☁️   Uploaded hotel_prices.xlsx → OneDrive/{folder}/")
    else:
        print(f"  ❌  OneDrive upload failed: {resp.status_code} {resp.text[:200]}")
        sys.exit(1)


# ── Main ──────────────────────────────────────────────────

async def main():
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        print("playwright not installed.")
        sys.exit(1)

    print(f"\n🏨  Hotel Price Scraper — {date.today()}")
    print(f"    {len(HOTELS)} hotels × {len(DATE_RANGES)} date ranges = "
          f"{len(HOTELS) * len(DATE_RANGES)} combinations\n")

    results: list[RoomResult] = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="en-GB",
            viewport={"width": 1400, "height": 900},
            extra_http_headers={"Accept-Language": "en-GB,en;q=0.9"},
        )
        page = await context.new_page()

        for hotel in HOTELS:
            for check_in, check_out in DATE_RANGES:
                result = await fetch_hotel(page, hotel, check_in, check_out)
                if result:
                    results.append(result)
                await asyncio.sleep(2)

        await browser.close()

    if not results:
        print("\n⚠️  No results scraped — skipping upload.")
        sys.exit(0)

    print(f"\n💾  Building workbook with {len(results)} new rows…")
    file_bytes = build_workbook(results)
    upload_to_onedrive(file_bytes)
    print(f"\n✅  Done. {len(results)} rows written.")


if __name__ == "__main__":
    asyncio.run(main())
